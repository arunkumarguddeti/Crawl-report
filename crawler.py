"""
crawler.py — Async Website Deep Crawler with Broken Link Detection
Generates a CSV data file + self-contained HTML dashboard report.

Configuration: edit the CONFIG block below, or pass env vars / CLI args.
Usage:
    python crawler.py
    BASE_URL=https://example.com python crawler.py
"""

import asyncio
import aiohttp
import csv
import json
import os
import re
import signal
import socket
import sys
import time
import logging
from collections import deque
from datetime import datetime, timezone, timedelta
from zoneinfo import ZoneInfo
from urllib.parse import urljoin, urlparse, urlunparse, urlencode, parse_qs
from urllib.robotparser import RobotFileParser
from pathlib import Path

# Eastern Time (handles EST/EDT automatically)
TZ_EASTERN = ZoneInfo("America/New_York")

def now_est() -> datetime:
    return datetime.now(TZ_EASTERN)

def fmt_duration(seconds: float) -> str:
    s = int(seconds)
    h, rem = divmod(s, 3600)
    m, sec = divmod(rem, 60)
    return f"{h}:{m:02d}:{sec:02d}"

# ─────────────────────────────────────────────────────────────
#  CONFIG  — edit these or override with environment variables
# ─────────────────────────────────────────────────────────────
CONFIG = {
    "BASE_URL":        os.getenv("BASE_URL",    "https://www.example.com"),
    "MAX_DEPTH":       int(os.getenv("MAX_DEPTH",   "10")),
    "MAX_PAGES":       int(os.getenv("MAX_PAGES",   "5000")),   # hard cap
    "CONCURRENCY":     int(os.getenv("CONCURRENCY", "10")),     # async workers
    "TIMEOUT":         int(os.getenv("TIMEOUT",     "10")),     # seconds per request
    "POLITE_DELAY":  float(os.getenv("POLITE_DELAY","0.3")),    # seconds between batches
    "RESPECT_ROBOTS":  os.getenv("RESPECT_ROBOTS", "true").lower() == "true",
    "OUTPUT_DIR":      os.getenv("OUTPUT_DIR",  "./reports"),
    "USER_AGENT":      "Mozilla/5.0 (compatible; SiteCrawler/1.0; +https://github.com/your-repo)",
    # Query params to strip when normalising URLs (avoid infinite pagination traps)
    # CRITICAL: do NOT add "page" — that strips paginated listing URLs like
    # /news?page=2, collapsing them to /news (already visited → dropped).
    # That single mistake caused crawls to stop at 116 pages instead of 3000+.
    "STRIP_PARAMS":    {"utm_source","utm_medium","utm_campaign","utm_content",
                        "utm_term","sessionid","PHPSESSID","sid","ref",
                        "search_api_fulltext","search_api_sort"},
    # File extensions to skip parsing (still HEAD-check, just don't crawl for more links)
    "SKIP_PARSE_EXTS": {".pdf",".zip",".docx",".xlsx",".pptx",".exe",".dmg",
                        ".mp4",".mp3",".avi",".mov",".jpg",".jpeg",".png",
                        ".gif",".svg",".webp",".ico",".woff",".woff2",".ttf"},
    # Schemes that are not HTTP — skip entirely
    "SKIP_SCHEMES":    {"mailto","tel","javascript","data","ftp","sms","callto"},

    # ── Queue explosion guard ──────────────────────────────────────────────
    # 20 000 is plenty for large sites; prevents Drupal facet/search runaway.
    "MAX_QUEUE":   int(os.getenv("MAX_QUEUE", "20000")),

    # URL path fragments that signal crawl-trap pages.
    # HEAD-checked as links (broken ones appear in report) but NOT crawled.
    #
    # REMOVED (were blocking real gov/Drupal content):
    #   /page/     Drupal path pagination: /news/page/2 is real content
    #   /archive/  gov sites have real /archive/year/ sections
    #   /author/   staff pages are real content
    #   /tag/ /tags/ /category/ /categories/  Drupal taxonomy listing pages
    #
    # ADDED Drupal machine-generated endpoints (not real pages):
    #   /media-library  thousands of facet URL permutations
    #   /search?        query-based search result pages
    #   search_api_page Drupal Search API pager params
    #   /views/ajax     Drupal Views AJAX JSON (not HTML)
    #   /jsonapi/       Drupal JSON:API data endpoint
    "TRAP_PATTERNS": set(
        os.getenv("TRAP_PATTERNS",
            "/feed/,/rss/,/wp-json/,/wp-admin/,"
            "/?s=,/?p=,/paged=,/?query=,/?q=,"
            "/media-library,/search?,search_api_page,"
            "/views/ajax,/jsonapi/"
        ).split(",")
    ),

    # ── Document extensions: HEAD-check only, never crawl ─────────────────
    "DOC_EXTS": {".pdf",".doc",".docx",".xls",".xlsx",".ppt",".pptx",
                ".odt",".ods",".odp",".zip",".tar",".gz",".7z",".rar",
                ".exe",".dmg",".pkg",".deb",".rpm"},

    # ── Rate-limit backoff ────────────────────────────────────────────────
    "RATE_LIMIT_BASE_WAIT": float(os.getenv("RATE_LIMIT_BASE_WAIT", "5.0")),
    "RATE_LIMIT_MAX_WAIT":  float(os.getenv("RATE_LIMIT_MAX_WAIT",  "120.0")),

    # ── Sitemap seeding ───────────────────────────────────────────────────
    "SEED_SITEMAP": os.getenv("SEED_SITEMAP", "true").lower() == "true",

    # ── Pagination depth guard ────────────────────────────────────────────
    # Set > 0 to block /news/page/N and ?page=N beyond this limit.
    # Default 0 = follow all pagination (safe when SEED_SITEMAP is on).
    "MAX_PAGINATION_PAGE": int(os.getenv("MAX_PAGINATION_PAGE", "0")),

    # ── Targeted page scan ────────────────────────────────────────────────
    # Comma-separated list of specific page URLs to scan.
    # When set, the full BFS crawl is skipped entirely — only these exact
    # pages are fetched and their links are checked.
    # Example: "https://site.com/about,https://site.com/contact"
    # Set via env var TARGET_PAGES or the GitHub Actions workflow input.
    "TARGET_PAGES": [
        u.strip() for u in os.getenv("TARGET_PAGES", "").split(",")
        if u.strip()
    ],

    # Optional label shown in the report header for targeted scans
    "SCAN_LABEL": os.getenv("SCAN_LABEL", ""),

    # ── Checkpoint / resume ───────────────────────────────────────────────────
    # CRAWL_TIMEOUT_SECS: stop after N seconds and save checkpoint for next run.
    # 0 = no limit (default for local/Mac use).
    # GitHub Actions sets this to 20700 (5h45m) via the workflow env.
    "CRAWL_TIMEOUT_SECS": int(os.getenv("CRAWL_TIMEOUT_SECS", "0")),
    "CHECKPOINT_IN":  os.getenv("CHECKPOINT_IN",  ""),   # path to resume from
    "CHECKPOINT_OUT": os.getenv("CHECKPOINT_OUT", ""),   # path to save to on timeout
}

# ─────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[logging.StreamHandler(sys.stdout)]
)
log = logging.getLogger("crawler")


# ──────────────────────────────────────────────
#  URL HELPERS
# ──────────────────────────────────────────────

def normalize(url: str, base: str = "") -> str | None:
    """
    Clean a URL:
    - Resolve relative URLs against `base`
    - Strip fragments (#section)
    - Lowercase scheme + host
    - Strip tracking/session query params
    - Return None for non-HTTP schemes or empty strings
    """
    if not url:
        return None
    url = url.strip()

    # Resolve relative URLs
    if base:
        url = urljoin(base, url)

    try:
        p = urlparse(url)
    except Exception:
        return None

    # Skip non-HTTP(S) schemes
    scheme = p.scheme.lower()
    if scheme in CONFIG["SKIP_SCHEMES"]:
        return None
    if scheme not in ("http", "https"):
        return None

    # Strip fragment, lowercase scheme+host, clean params
    params = parse_qs(p.query, keep_blank_values=True)
    cleaned = {k: v for k, v in params.items()
               if k not in CONFIG["STRIP_PARAMS"]}
    clean_query = urlencode(cleaned, doseq=True)

    clean = urlunparse((
        scheme,
        p.netloc.lower(),
        p.path.rstrip("/") or "/",
        p.params,
        clean_query,
        ""  # no fragment
    ))
    return clean


def _bare_domain(netloc: str) -> str:
    """Strip www. so www.site.gov and site.gov are treated as the same site."""
    return netloc.lower().removeprefix("www.")


def is_same_domain(url: str, base: str) -> bool:
    """True if url is on the same domain as base (www/non-www treated equally)."""
    return _bare_domain(urlparse(url).netloc) == _bare_domain(urlparse(base).netloc)


def should_skip_parse(url: str) -> bool:
    """True if URL points to a binary/document file — should not parse for links."""
    path = urlparse(url).path.lower()
    all_skip = CONFIG["SKIP_PARSE_EXTS"] | CONFIG.get("DOC_EXTS", set())
    return any(path.endswith(ext) for ext in all_skip)


def is_doc_url(url: str) -> bool:
    """
    True for document/file URLs: HEAD-check to detect 404 broken file links
    but NEVER add to the BFS crawl queue.
    Covers Drupal /files/ paths and DOC_EXTS extensions.
    """
    path = urlparse(url).path.lower()
    if "/files/" in path:
        return True
    return any(path.endswith(ext) for ext in CONFIG.get("DOC_EXTS", set()))


# ──────────────────────────────────────────────
#  ROBOTS.TXT
# ──────────────────────────────────────────────

def load_robots(base_url: str) -> RobotFileParser | None:
    if not CONFIG["RESPECT_ROBOTS"]:
        return None
    rp = RobotFileParser()
    rp.set_url(base_url.rstrip("/") + "/robots.txt")
    try:
        rp.read()
        log.info("robots.txt loaded from %s", base_url)
    except Exception as e:
        log.warning("Could not read robots.txt: %s", e)
        return None
    return rp


def robots_allow(rp: RobotFileParser | None, url: str) -> bool:
    if rp is None:
        return True
    return rp.can_fetch(CONFIG["USER_AGENT"], url)


# ──────────────────────────────────────────────
#  EFFORT CLASSIFICATION
# ──────────────────────────────────────────────

def effort_level(status) -> str:
    s = str(status)
    if s == "200":                           return "None"
    if s.startswith("3"):                    return "Low"
    if s in ("401", "403", "429"):           return "Low-Medium"
    if s in ("404", "410", "451"):           return "High"
    if s.startswith("5"):                    return "Medium"
    if "Timeout" in s or "SSL" in s:         return "Medium"
    if "Error" in s or "Connection" in s:    return "High"
    return "Medium"

def status_category(status) -> str:
    s = str(status)
    if s == "200":                return "OK"
    if s.startswith("2"):         return "2xx Other"
    if s.startswith("3"):         return "Redirect"
    if s in ("404", "410"):       return "Not Found"
    if s.startswith("4"):         return "4xx Client Error"
    if s.startswith("5"):         return "5xx Server Error"
    if "Timeout" in s:            return "Timeout"
    return "Error"


# ──────────────────────────────────────────────
#  ASYNC HTTP HELPERS
# ──────────────────────────────────────────────

# ── Rate-limit shared state ───────────────────────────────────────────────
_rl_wait: float = 0.0   # extra inter-batch sleep (0 = not rate-limited)
_rl_hits: int   = 0     # consecutive 429 count


async def _handle_429(url: str, retry_after: str = "") -> None:
    """Exponential back-off when server returns 429 Too Many Requests."""
    global _rl_wait, _rl_hits
    _rl_hits += 1
    if retry_after:
        try:
            wait = float(retry_after)
        except ValueError:
            wait = CONFIG.get("RATE_LIMIT_BASE_WAIT", 5.0)
    else:
        base  = CONFIG.get("RATE_LIMIT_BASE_WAIT", 5.0)
        limit = CONFIG.get("RATE_LIMIT_MAX_WAIT", 120.0)
        wait  = min(base * (2 ** (_rl_hits - 1)), limit)
    _rl_wait = wait
    log.warning(
        "\U0001f6a6 429 on %s — backing off %.0fs (hit #%d). "
        "Stop any other parallel crawl instances!",
        url[:70], wait, _rl_hits)
    await asyncio.sleep(wait)


def _reset_rl() -> None:
    """Reset rate-limit state after a successful response."""
    global _rl_wait, _rl_hits
    if _rl_hits > 0:
        log.info("\u2705 Rate limit cleared — resuming normal speed.")
    _rl_wait = 0.0
    _rl_hits = 0


async def fetch_page_html(session: aiohttp.ClientSession, url: str) -> tuple[int, str, str]:
    """
    GET a page. Returns (status_code, html_text, final_url_after_redirects).
    Handles 429 with exponential back-off. Retries up to 3× on timeout.
    """
    headers = {"User-Agent": CONFIG["USER_AGENT"]}
    timeout = aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"])

    for attempt in range(1, 4):
        try:
            async with session.get(url, headers=headers, timeout=timeout,
                                   allow_redirects=True, ssl=False) as resp:
                if resp.status == 429:
                    await _handle_429(url, resp.headers.get("Retry-After", ""))
                    continue
                _reset_rl()
                try:
                    html = await resp.text(errors="ignore")
                except Exception:
                    html = ""
                return resp.status, html, str(resp.url)
        except asyncio.TimeoutError:
            if attempt < 3:
                await asyncio.sleep(attempt * 2)
                continue
            return "Timeout", "", url
        except aiohttp.ClientSSLError:
            return "SSL Error", "", url
        except aiohttp.ClientConnectorError:
            return "Connection Error", "", url
        except Exception as e:
            return f"Error: {type(e).__name__}", "", url
    return "Timeout", "", url


async def check_link_status(session: aiohttp.ClientSession,
                             url: str,
                             link_cache: dict) -> tuple:
    """
    HEAD (then GET fallback) a URL.
    Returns (status, final_url, load_ms) — cached to avoid duplicate requests.
    """
    if url in link_cache:
        return link_cache[url]

    headers = {"User-Agent": CONFIG["USER_AGENT"]}
    timeout = aiohttp.ClientTimeout(total=CONFIG["TIMEOUT"])
    t0 = time.monotonic()
    for _attempt in range(3):
        try:
            async with session.head(url, headers=headers, timeout=timeout,
                                    allow_redirects=True, ssl=False) as resp:
                load_ms = round((time.monotonic() - t0) * 1000)
                if resp.status == 429:
                    await _handle_429(url, resp.headers.get("Retry-After", ""))
                    t0 = time.monotonic()
                    continue
                # 405 = HEAD not allowed; 403 = WAF blocks HEAD but GET works.
                # Both trigger a GET fallback for the real status code.
                if resp.status in (405, 403):
                    raise aiohttp.ClientResponseError(
                        resp.request_info, resp.history, status=resp.status)
                _reset_rl()
                result = (resp.status, str(resp.url), load_ms)
                break
        except aiohttp.ClientResponseError:
            t0 = time.monotonic()
            try:
                async with session.get(url, headers=headers, timeout=timeout,
                                       allow_redirects=True, ssl=False) as resp:
                    load_ms = round((time.monotonic() - t0) * 1000)
                    if resp.status == 429:
                        await _handle_429(url, resp.headers.get("Retry-After", ""))
                        t0 = time.monotonic()
                        continue
                    _reset_rl()
                    result = (resp.status, str(resp.url), load_ms)
                    break
            except asyncio.TimeoutError:
                result = ("Timeout", url, -1); break
            except Exception as e:
                result = (f"Error: {type(e).__name__}", url, -1); break
        except asyncio.TimeoutError:
            result = ("Timeout", url, -1); break
        except aiohttp.ClientSSLError:
            result = ("SSL Error", url, -1); break
        except aiohttp.ClientConnectorError:
            result = ("Connection Error", url, -1); break
        except Exception as e:
            result = (f"Error: {type(e).__name__}", url, -1); break
    else:
        result = ("Timeout", url, -1)

    link_cache[url] = result
    return result


# ──────────────────────────────────────────────
#  HTML PARSER  (no BeautifulSoup needed — stdlib)
# ──────────────────────────────────────────────

def extract_links(html: str, page_url: str) -> list[tuple[str, str]]:
    """
    Extract all <a href> links + anchor text from HTML.
    Returns list of (absolute_url, anchor_text) tuples.
    Falls back to BeautifulSoup if available, else uses stdlib html.parser.
    """
    try:
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(html, "lxml")
        links = []
        for a in soup.find_all("a", href=True):
            href = a.get("href", "").strip()
            text = a.get_text(separator=" ", strip=True)[:200]
            norm = normalize(href, page_url)
            if norm:
                links.append((norm, text))
        return links
    except ImportError:
        pass

    # Stdlib fallback using html.parser
    from html.parser import HTMLParser

    class LinkParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.links = []
            self._in_a = False
            self._current_href = None
            self._current_text = []

        def handle_starttag(self, tag, attrs):
            if tag == "a":
                self._in_a = True
                d = dict(attrs)
                href = d.get("href", "").strip()
                self._current_href = normalize(href, page_url)
                self._current_text = []

        def handle_data(self, data):
            if self._in_a:
                self._current_text.append(data.strip())

        def handle_endtag(self, tag):
            if tag == "a" and self._in_a:
                if self._current_href:
                    text = " ".join(t for t in self._current_text if t)[:200]
                    self.links.append((self._current_href, text))
                self._in_a = False
                self._current_href = None
                self._current_text = []

    parser = LinkParser()
    try:
        parser.feed(html)
    except Exception:
        pass
    return parser.links


# ──────────────────────────────────────────────
#  TARGETED CRAWL  — scan a specific list of pages only
#  No BFS. No following links. Just fetch each given page,
#  check every link on it, and return results.
# ──────────────────────────────────────────────

async def targeted_crawl(target_urls: list[str]) -> list[dict]:
    """
    Fetch each URL in target_urls, check all their outbound links,
    and return results — without crawling further.
    """
    base_url   = CONFIG["BASE_URL"].rstrip("/")
    sem        = asyncio.Semaphore(CONFIG["CONCURRENCY"])
    link_cache: dict = {}
    results:    list[dict] = []

    log.info("TARGETED MODE — scanning %d specific page(s):", len(target_urls))
    for u in target_urls:
        log.info("  → %s", u)

    connector = aiohttp.TCPConnector(
        limit=CONFIG["CONCURRENCY"] + 5,
        ssl=False,
        force_close=False,
        enable_cleanup_closed=True,
    )

    async with aiohttp.ClientSession(connector=connector) as session:

        async def scan_page(page_url: str):
            log.info("Scanning page: %s", page_url)
            async with sem:
                status, html, final_url = await fetch_page_html(session, page_url)

            if str(status) != "200":
                results.append({
                    "page_url":  page_url,
                    "link_url":  page_url,
                    "link_text": "(page itself)",
                    "link_type": "Page",
                    "status":    status,
                    "final_url": final_url,
                    "load_ms":   -1,
                    "depth":     0,
                    "effort":    effort_level(status),
                    "category":  status_category(status),
                    "timestamp": now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                })
                return

            if should_skip_parse(page_url):
                return

            page_links = extract_links(html, page_url)
            if not page_links:
                log.info("  No links found on %s", page_url)
                return

            log.info("  Found %d links on %s", len(page_links), page_url)

            async def check_one(link_url: str, link_text: str):
                async with sem:
                    lnk_status, lnk_final, lnk_load_ms = await check_link_status(
                        session, link_url, link_cache)
                link_type = ("Internal"
                             if is_same_domain(link_url, base_url)
                             else "External")
                results.append({
                    "page_url":  page_url,
                    "link_url":  link_url,
                    "link_text": link_text,
                    "link_type": link_type,
                    "status":    lnk_status,
                    "final_url": lnk_final,
                    "load_ms":   lnk_load_ms,
                    "depth":     0,
                    "effort":    effort_level(lnk_status),
                    "category":  status_category(lnk_status),
                    "timestamp": now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                })

            await asyncio.gather(*[check_one(lu, lt) for lu, lt in page_links])

        # Scan all target pages concurrently (respecting semaphore)
        await asyncio.gather(*[scan_page(u) for u in target_urls])

    log.info("Targeted scan complete. %d link records collected.", len(results))
    return results


# ──────────────────────────────────────────────
#  MAIN CRAWL  (BFS + async 10-worker semaphore)
# ──────────────────────────────────────────────

def is_trap_url(url: str) -> bool:
    """
    True if the URL matches a known crawl-trap pattern.
    These are HEAD-checked as links (so broken ones appear in the report)
    but are NOT added to the BFS crawl queue.
    """
    lower = url.lower()
    if any(pat.strip() and pat.strip() in lower for pat in CONFIG["TRAP_PATTERNS"]):
        return True
    # Pagination depth guard: only active when MAX_PAGINATION_PAGE > 0.
    # Prevents following infinite archive pages (page 500, page 501 …)
    # while still allowing /news/page/2, /events/page/3 by default.
    max_pag = CONFIG.get("MAX_PAGINATION_PAGE", 0)
    if max_pag > 0:
        m = re.search(r"/page/(\d+)", lower)
        if m and int(m.group(1)) > max_pag:
            return True
        m2 = re.search(r"[?&]page=(\d+)", lower)
        if m2 and int(m2.group(1)) > max_pag:
            return True
    return False


def safe_enqueue(queue: deque, seen: set, url: str, depth: int) -> bool:
    """
    Add a URL to the BFS crawl queue only when ALL conditions are met:
      1. Not already seen (visited or queued)
      2. Depth within MAX_DEPTH (rejected without adding to seen, so shallower
         re-discovery at a later crawl step still works)
      3. Queue below MAX_QUEUE cap
      4. Not a trap URL pattern
      5. Not a binary file extension
      6. Not a document / /files/ path (those are HEAD-checked as links only)
    """
    if url in seen:
        return False
    if depth > CONFIG["MAX_DEPTH"]:
        return False          # NOT added to seen — allow shallower re-discovery
    if len(queue) >= CONFIG["MAX_QUEUE"]:
        log.warning("MAX_QUEUE cap (%d) reached — URL dropped: %s",
                    CONFIG["MAX_QUEUE"], url[:80])
        return False
    if is_trap_url(url):
        return False
    if should_skip_parse(url):
        return False
    # /files/ paths and document extensions: verify as links (HEAD-check)
    # but never BFS-crawl them for more links.
    if is_doc_url(url):
        return False
    seen.add(url)
    queue.append((url, depth))
    return True



# ──────────────────────────────────────────────
#  DNS / IP RESOLUTION
# ──────────────────────────────────────────────

_dns_cache: dict[str, str] = {}   # hostname → resolved IP


def _resolve_sync(hostname: str) -> str:
    try:
        infos = socket.getaddrinfo(hostname, None)
        ipv4 = [i[4][0] for i in infos if i[0] == socket.AF_INET]
        ipv6 = [i[4][0] for i in infos if i[0] == socket.AF_INET6]
        return ipv4[0] if ipv4 else (ipv6[0] if ipv6 else "N/A")
    except Exception:
        return "N/A"


async def resolve_ip(hostname: str) -> str:
    """Resolve hostname → IP, cached per hostname."""
    if hostname in _dns_cache:
        return _dns_cache[hostname]
    loop = asyncio.get_event_loop()
    ip = await loop.run_in_executor(None, _resolve_sync, hostname)
    _dns_cache[hostname] = ip
    return ip


async def enrich_with_ips(results: list[dict]) -> None:
    """
    Resolve IPs for two things per result row:
      - page_ip : IP of the parent page (the page being crawled)
      - link_ip : IP of the link URL (could be external)
    Both are added as fields and appear in the HTML report.
    One DNS lookup per unique hostname, cached and resolved concurrently.
    """
    from urllib.parse import urlparse as _up
    hostnames = set()
    for r in results:
        h1 = _up(r.get("page_url", "")).hostname or ""
        h2 = _up(r.get("link_url", "")).hostname or ""
        if h1: hostnames.add(h1)
        if h2: hostnames.add(h2)
    hostnames.discard("")
    log.info("\U0001f50d Resolving IPs for %d unique hostnames…", len(hostnames))
    await asyncio.gather(*[resolve_ip(h) for h in hostnames])
    resolved = sum(1 for h in hostnames if _dns_cache.get(h, "N/A") != "N/A")
    log.info("\U0001f50d IP resolution: %d/%d resolved", resolved, len(hostnames))
    for r in results:
        ph = _up(r.get("page_url", "")).hostname or ""
        lh = _up(r.get("link_url", "")).hostname or ""
        r["page_ip"] = _dns_cache.get(ph, "N/A")
        r["link_ip"] = _dns_cache.get(lh, "N/A")


# ──────────────────────────────────────────────
#  SITEMAP SEEDER
# ──────────────────────────────────────────────

async def _seed_from_sitemap(base_url: str, queue: "deque",
                              seen: set, session: "aiohttp.ClientSession") -> int:
    """
    Discover and parse XML sitemaps, seeding all internal URLs into the queue.
    Tries robots.txt first for Sitemap: directives, then common sitemap paths.
    Handles nested sitemap indexes. Sitemap URLs bypass trap filters because
    they are explicitly declared by the site as canonical content pages.
    """
    added    = 0
    sm_queue: list[str] = []
    visited  : set[str] = set()
    headers  = {"User-Agent": CONFIG["USER_AGENT"]}
    timeout  = aiohttp.ClientTimeout(total=20)
    base     = base_url.rstrip("/")

    # Step 1: robots.txt for Sitemap: directives
    try:
        async with session.get(f"{base}/robots.txt", headers=headers,
                               timeout=timeout, ssl=False) as resp:
            if resp.status == 200:
                for line in (await resp.text(errors="ignore")).splitlines():
                    if line.strip().lower().startswith("sitemap:"):
                        sm_url = line.split(":", 1)[1].strip()
                        if sm_url and sm_url not in visited:
                            sm_queue.append(sm_url)
                            log.info("\U0001f5fa  Sitemap from robots.txt: %s", sm_url)
    except Exception as e:
        log.debug("robots.txt sitemap check failed: %s", e)

    # Step 2: common paths if robots.txt had none
    if not sm_queue:
        for path in ["/sitemap.xml", "/sitemap_index.xml", "/sitemap/",
                     "/sitemaps/1", "/sitemap-index.xml"]:
            sm_queue.append(base + path)

    # Step 3: fetch and parse
    while sm_queue:
        sm_url = sm_queue.pop(0)
        if sm_url in visited:
            continue
        visited.add(sm_url)
        try:
            async with session.get(sm_url, headers=headers,
                                   timeout=timeout, ssl=False) as resp:
                if resp.status != 200:
                    continue
                ct = resp.headers.get("Content-Type", "")
                if "html" in ct.lower():
                    continue  # got an HTML error page, not XML
                text = await resp.text(errors="ignore")
        except Exception as e:
            log.debug("Sitemap fetch failed %s: %s", sm_url, e)
            continue

        # Child sitemaps
        for loc in re.findall(r"<sitemap>\s*<loc>\s*(.*?)\s*</loc>", text, re.DOTALL):
            if loc not in visited:
                sm_queue.append(loc.strip())

        # Page URLs — add directly, bypassing trap filters
        for loc in re.findall(r"<url>\s*<loc>\s*(.*?)\s*</loc>", text, re.DOTALL):
            loc  = loc.strip()
            norm = normalize(loc)
            if not norm or not is_same_domain(norm, base_url):
                continue
            if norm in seen or is_doc_url(norm):
                continue
            seen.add(norm)
            queue.append((norm, 0))
            added += 1

    if added:
        log.info("\U0001f5fa  Sitemap seeding: %d URLs added to queue", added)
    else:
        log.warning("\U0001f5fa  No sitemap found — relying on link discovery only")
    return added


async def crawl() -> tuple[list[dict], bool]:
    """
    Returns (results, checkpoint_saved).
    Ctrl-C / SIGTERM triggers graceful shutdown: the current batch finishes,
    then full reports are written from whatever was collected so far.
    """
    base_url = CONFIG["BASE_URL"].rstrip("/")
    sem      = asyncio.Semaphore(CONFIG["CONCURRENCY"])

    # ── Graceful shutdown (Ctrl-C / SIGTERM) ──────────────────────────────
    _shutdown = asyncio.Event()

    def _on_signal(sig_name: str):
        if not _shutdown.is_set():
            log.warning("⚠️  %s — finishing batch then writing partial report…",
                        sig_name)
            _shutdown.set()

    _loop = asyncio.get_event_loop()
    for _s in (signal.SIGINT, signal.SIGTERM):
        try:
            _loop.add_signal_handler(_s, lambda n=_s.name: _on_signal(n))
        except (NotImplementedError, ValueError):
            pass  # Windows

    visited_pages:   set[str]   = set()  # pages whose HTML fetch was started
    completed_pages: set[str]   = set()  # pages fully processed (all links checked)
    seen:            set[str]   = set()  # visited + queued (dedup guard)
    link_cache:      dict       = {}     # url -> (status, final_url, ms) cache
    results:         list[dict] = []
    trap_skipped:    int        = 0
    prior_run_count: int        = 0      # number of chained runs before this one

    # ── Load checkpoint if resuming ───────────────────────────────────────────
    checkpoint_in = CONFIG["CHECKPOINT_IN"]
    if checkpoint_in and Path(checkpoint_in).exists():
        log.info("📂 Loading checkpoint: %s", checkpoint_in)
        with open(checkpoint_in) as f:
            ckpt = json.load(f)
        visited_pages   = set(ckpt["visited_pages"])
        completed_pages = set(ckpt["completed_pages"])
        seen            = set(ckpt["seen"])
        link_cache      = {k: tuple(v) for k, v in ckpt["link_cache"].items()}
        results         = ckpt["results"]
        trap_skipped    = ckpt["trap_skipped"]
        prior_run_count = ckpt.get("run_count", 1)
        queue_data      = ckpt["queue"]
        # Pages that were in-progress (visited but not completed) go back to queue
        interrupted = visited_pages - completed_pages
        if interrupted:
            log.info("♻️  Re-queuing %d interrupted pages (were in-progress at cutoff)",
                     len(interrupted))
            for url in sorted(interrupted):
                seen.discard(url)
                visited_pages.discard(url)
        queue: deque[tuple[str, int]] = deque(
            (url, depth) for url, depth in queue_data
        )
        log.info("✅ Checkpoint loaded: %d visited, %d completed, %d queued, "
                 "%d cached links, %d results so far",
                 len(visited_pages), len(completed_pages),
                 len(queue), len(link_cache), len(results))
    else:
        rp = load_robots(base_url)
        start = normalize(base_url)
        if not start:
            log.error("BASE_URL '%s' is invalid.", base_url)
            return [], False
        queue: deque[tuple[str, int]] = deque()
        seen.add(start)
        queue.append((start, 0))

    rp = load_robots(base_url)
    crawl_start    = time.time()
    timeout_secs   = CONFIG["CRAWL_TIMEOUT_SECS"]
    checkpoint_saved = False

    connector = aiohttp.TCPConnector(
        limit=CONFIG["CONCURRENCY"] + 5,
        ssl=False,
        force_close=False,
        enable_cleanup_closed=True,
    )

    async with aiohttp.ClientSession(connector=connector) as session:

        # ── Sitemap seeding on fresh crawl ────────────────────────────────
        if CONFIG.get("SEED_SITEMAP") and not CONFIG.get("CHECKPOINT_IN", ""):
            await _seed_from_sitemap(base_url, queue, seen, session)

        while queue:
            # ── Graceful shutdown: Ctrl-C / SIGTERM ───────────────────────
            if _shutdown.is_set():
                log.warning("Shutdown requested — writing partial report "
                            "(%d results collected).", len(results))
                break

            # ── Hard stop: MAX_PAGES reached ──────────────────────────────
            if CONFIG["MAX_PAGES"] > 0 and len(visited_pages) >= CONFIG["MAX_PAGES"]:
                log.warning("MAX_PAGES cap (%d) reached — crawl complete.",
                            CONFIG["MAX_PAGES"])
                break

            # ── Timeout check: save checkpoint and pause for next run ──────
            if timeout_secs > 0 and (time.time() - crawl_start) >= timeout_secs:
                elapsed_min = (time.time() - crawl_start) / 60
                log.warning(
                    "⏱️  Crawl timeout reached after %.1f min — saving checkpoint "
                    "(%d pages done, %d queued). Next run will resume from here.",
                    elapsed_min, len(completed_pages), len(queue))
                checkpoint_out = CONFIG["CHECKPOINT_OUT"]
                if checkpoint_out:
                    ckpt = {
                        "run_count":       prior_run_count + 1,
                        "base_url":        base_url,
                        "visited_pages":   list(visited_pages),
                        "completed_pages": list(completed_pages),
                        "seen":            list(seen),
                        "queue":           list(queue),
                        # link_cache values are tuples — serialise as lists
                        "link_cache":      {k: list(v) for k, v in link_cache.items()},
                        "results":         results,
                        "trap_skipped":    trap_skipped,
                        "saved_at":        now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                    }
                    Path(checkpoint_out).parent.mkdir(parents=True, exist_ok=True)
                    with open(checkpoint_out, "w") as f:
                        json.dump(ckpt, f)
                    log.info("💾 Checkpoint saved → %s  (%.1f MB)",
                             checkpoint_out,
                             Path(checkpoint_out).stat().st_size / 1_048_576)
                    checkpoint_saved = True
                break

            # ── Build next batch ──────────────────────────────────────────
            batch: list[tuple[str, int]] = []
            while queue and len(batch) < CONFIG["CONCURRENCY"]:
                url, depth = queue.popleft()
                if url in visited_pages:
                    continue
                if depth > CONFIG["MAX_DEPTH"]:
                    continue
                if not robots_allow(rp, url):
                    log.debug("robots.txt disallows: %s", url)
                    continue
                visited_pages.add(url)
                batch.append((url, depth))

            if not batch:
                break

            log.info("Crawling batch of %d pages  |  visited: %d  |"
                     "  queue: %d  |  trap-skipped: %d",
                     len(batch), len(visited_pages),
                     len(queue), trap_skipped)

            # ── Process each page in the batch concurrently ───────────────
            async def process_page(page_url: str, depth: int):
                nonlocal trap_skipped

                async with sem:
                    status, html, final_url = await fetch_page_html(
                        session, page_url)

                # If the page itself is broken, record it and stop
                if str(status) != "200":
                    results.append({
                        "page_url":  page_url,
                        "link_url":  page_url,
                        "link_text": "(page itself)",
                        "link_type": "Page",
                        "status":    status,
                        "final_url": final_url,
                        "load_ms":   -1,
                        "depth":     depth,
                        "effort":    effort_level(status),
                        "category":  status_category(status),
                        "timestamp": now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                    })
                    return

                if should_skip_parse(page_url):
                    return

                page_links = extract_links(html, page_url)
                if not page_links:
                    return

                # ── Check every link on this page (concurrent, cached) ────
                async def check_one(link_url: str, link_text: str):
                    nonlocal trap_skipped
                    async with sem:
                        lnk_status, lnk_final, lnk_load_ms = await check_link_status(
                            session, link_url, link_cache)

                    link_type = ("Internal"
                                 if is_same_domain(link_url, base_url)
                                 else "External")
                    results.append({
                        "page_url":  page_url,
                        "link_url":  link_url,
                        "link_text": link_text,
                        "link_type": link_type,
                        "status":    lnk_status,
                        "final_url": lnk_final,
                        "load_ms":   lnk_load_ms,
                        "depth":     depth,
                        "effort":    effort_level(lnk_status),
                        "category":  status_category(lnk_status),
                        "timestamp": now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                    })

                    # ── Decide whether to crawl this link ─────────────────
                    if link_type != "Internal":
                        return  # never crawl external domains
                    if CONFIG["MAX_PAGES"] > 0 and len(visited_pages) >= CONFIG["MAX_PAGES"]:
                        return  # hard page cap
                    # Only skip crawling if DEFINITIVELY broken.
                    # 404/410/451 = does not exist → skip.
                    # Timeout/Error/5xx on HEAD are transient — still enqueue;
                    # the actual GET when we crawl the page reveals true status.
                    s = str(lnk_status)
                    if s in ("404", "410", "451"):
                        return

                    # Trap-pattern check with counter for reporting
                    if is_trap_url(link_url):
                        trap_skipped += 1
                        return

                    # Enqueue immediately — safe_enqueue deduplicates via seen set
                    safe_enqueue(queue, seen, link_url, depth + 1)

                await asyncio.gather(
                    *[check_one(lu, lt) for lu, lt in page_links])

            await asyncio.gather(*[process_page(u, d) for u, d in batch])

            # Mark all pages in this batch as fully completed
            for page_url, _ in batch:
                completed_pages.add(page_url)

            # ── Adaptive polite delay ─────────────────────────────────────
            _delay = max(CONFIG["POLITE_DELAY"], _rl_wait)
            if _delay > 0:
                if _rl_wait > CONFIG["POLITE_DELAY"]:
                    log.info("⏳ Rate-limit cooldown: %.1fs between batches",
                             _delay)
                await asyncio.sleep(_delay)

    log.info(
        "Crawl complete. visited=%d  completed=%d  results=%d  "
        "trap-skipped=%d  queue-remaining=%d",
        len(visited_pages), len(completed_pages),
        len(results), trap_skipped, len(queue))
    return results, checkpoint_saved


# ──────────────────────────────────────────────
#  CSV OUTPUT
# ──────────────────────────────────────────────

FIELDS = ["page_url","link_url","link_text","link_type",
          "status","final_url","load_ms","depth","effort","category",
          "page_ip","link_ip","timestamp"]

def write_csv(results: list[dict], path: str):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)
    log.info("CSV written → %s  (%d rows)", path, len(results))


# ──────────────────────────────────────────────
#  HTML REPORT GENERATION
#  Data stored as JS JSON — only PAGE_SIZE rows in DOM at once.
#  All filtering/sorting runs on the JS array (never touches 5000+ DOM nodes).
#  Excel export uses SheetJS (CDN) — no server needed, works offline.
# ──────────────────────────────────────────────

def build_html_report(results: list[dict], csv_path: str, elapsed: float,
                      scan_mode: str = "full", target_pages: list = None) -> str:

    PAGE_SIZE = 100

    total_links   = len(results)
    pages_set     = {r["page_url"] for r in results}
    total_pages   = len(pages_set)
    broken        = sum(1 for r in results if str(r["status"]) in ("404","410","451"))
    redirects     = sum(1 for r in results if str(r["status"]).startswith("3"))
    server_errors = sum(1 for r in results if str(r["status"]).startswith("5"))
    ok_links      = sum(1 for r in results if str(r["status"]) == "200")
    errors        = sum(1 for r in results if "Error" in str(r["status"]) or "Timeout" in str(r["status"]))
    external      = sum(1 for r in results if r.get("link_type") == "External")
    internal      = sum(1 for r in results if r.get("link_type") == "Internal")

    cat_counts: dict[str, int] = {}
    for r in results:
        cat = r.get("category", "Unknown")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    broken_by_page: dict[str, int] = {}
    for r in results:
        s = str(r["status"])
        if s in ("404","410","451") or "Error" in s or "Timeout" in s:
            broken_by_page[r["page_url"]] = broken_by_page.get(r["page_url"], 0) + 1
    top_broken = sorted(broken_by_page.items(), key=lambda x: -x[1])[:10]

    chart_labels = json.dumps(list(cat_counts.keys()))
    chart_values = json.dumps(list(cat_counts.values()))
    chart_colors = json.dumps([
        "#22c55e","#16a34a","#f59e0b","#ef4444","#dc2626",
        "#3b82f6","#8b5cf6","#64748b","#0ea5e9","#f97316"
    ][:len(cat_counts)])
    bar_labels = json.dumps([p[:60]+"…" if len(p) > 60 else p for p,_ in top_broken])
    bar_values = json.dumps([v for _,v in top_broken])

    run_date   = now_est().strftime("%Y-%m-%d %H:%M:%S %Z")
    run_dur    = fmt_duration(elapsed)
    # In targeted mode, derive display URL from first target page if BASE_URL is still default
    _default_url = "https://www.example.com"
    if scan_mode == "targeted" and target_pages and CONFIG["BASE_URL"].rstrip("/") == _default_url.rstrip("/"):
        from urllib.parse import urlparse as _up
        _p = _up(target_pages[0])
        base_url = f"{_p.scheme}://{_p.netloc}"
    else:
        base_url = CONFIG["BASE_URL"]
    scan_label = CONFIG.get("SCAN_LABEL", "")
    mode_badge = (
        f'<span style="background:#7c3aed;color:#fff;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700;margin-left:10px">🎯 TARGETED SCAN</span>'
        if scan_mode == "targeted" else
        f'<span style="background:#0369a1;color:#fff;padding:2px 10px;border-radius:12px;font-size:12px;font-weight:700;margin-left:10px">🌐 FULL CRAWL</span>'
    )
    target_count = len(target_pages) if target_pages else 0
    # Build targeted pages info block for HTML
    if scan_mode == "targeted" and target_pages:
        tp_items = "".join(
            f'<li><a href="{u}" target="_blank" style="color:#93c5fd">{u}</a></li>'
            for u in target_pages
        )
        label_html = (
            '<b style="color:#94a3b8">Label: </b>'
            '<span style="color:#e2e8f0">' + scan_label + '</span><br><br>'
        ) if scan_label else ""
        targeted_info_html = (
            '<div style="background:#1e293b;border:1px solid #7c3aed;' +
            'border-radius:10px;padding:14px 18px;margin-bottom:18px">' +
            '<div style="font-weight:700;color:#c4b5fd;margin-bottom:8px">' +
            f"\U0001f3af Targeted Scan \u2014 {target_count} page(s) scanned" +
            '</div>' +
            label_html +
            '<ul style="list-style:none;display:flex;flex-direction:column;gap:4px;padding:0">' +
            tp_items +
            '</ul></div>'
        )

    else:
        targeted_info_html = ""

    def row_class(status):
        s = str(status)
        if s == "200":         return "ok"
        if s.startswith("3"):  return "redirect"
        if s in ("404","410"): return "broken"
        if s.startswith("4"):  return "warn"
        if s.startswith("5"):  return "servererr"
        return "unkn"

    def load_display(ms):
        if ms is None or ms == -1:
            return "—"
        if ms < 1000:
            return f"{ms} ms"
        return f"{ms/1000:.1f}s"

    # Embed ALL rows in HTML including 200 OK links.
    # 200 OK external links are important: they are valid today but could break
    # if the external site changes its URL structure without setting up redirects.
    # Users can filter them in/out via the "Show 200 OK" toggle in the report.
    js_rows = []
    for r in results:
        s = str(r.get("status",""))
        lm = r.get("load_ms", -1)
        js_rows.append({
            "pu": r["page_url"],
            "lu": r["link_url"],
            "lt": r.get("link_text","")[:120],
            "tp": r.get("link_type",""),
            "st": s,
            "fu": r.get("final_url",""),
            "lm": lm if lm is not None else -1,
            "dp": r.get("depth",""),
            "ef": r.get("effort",""),
            "ca": r.get("category",""),
            "pi": r.get("page_ip","N/A"),
            "li": r.get("link_ip","N/A"),
            "rc": row_class(r["status"]),
            "ts": str(r.get("timestamp",""))[:22],
        })
    # Web-safe JSON: escape < and > so NO HTML tag (</script>, </div>, etc.)
    # can ever break the script block, regardless of what URLs/text are crawled.
    # \u003c = <  |  \u003e = >  — valid JSON, browsers decode transparently.
    # Base64-encode the JSON so NO character in any crawled URL, anchor text,
    # or page content can ever break the HTML/JS parser.
    # atob() in JS decodes it back — base64 only contains [A-Za-z0-9+/=].
    import base64 as _b64
    all_data_b64 = _b64.b64encode(
        json.dumps(js_rows, ensure_ascii=False).encode("utf-8")
    ).decode("ascii")

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Broken Link Report — {base_url}</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.sheetjs.com/xlsx-0.20.1/package/dist/xlsx.full.min.js"></script>
<style>
  :root{{
    --bg:#0f172a;--surface:#1e293b;--surface2:#263348;
    --border:#334155;--text:#e2e8f0;--muted:#94a3b8;
    --green:#22c55e;--yellow:#f59e0b;--red:#ef4444;
    --blue:#3b82f6;--purple:#8b5cf6;--orange:#f97316;
    --radius:10px;
  }}
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{background:var(--bg);color:var(--text);font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;font-size:14px}}
  a{{color:#60a5fa;text-decoration:none}} a:hover{{text-decoration:underline}}

  .header{{background:var(--surface);border-bottom:1px solid var(--border);padding:16px 28px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}}
  .header h1{{font-size:1.25rem;font-weight:700}} .header h1 span{{color:var(--blue)}}
  .meta{{color:var(--muted);font-size:12px;display:flex;gap:16px;flex-wrap:wrap;margin-top:4px}} .meta b{{color:var(--text)}}
  .hdr-btns{{display:flex;gap:8px;flex-wrap:wrap}}

  .container{{max-width:1600px;margin:0 auto;padding:20px 28px}}

  .cards{{display:grid;grid-template-columns:repeat(auto-fit,minmax(148px,1fr));gap:12px;margin-bottom:22px}}
  .card{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:14px 16px;cursor:pointer;transition:border-color .15s}}
  .card:hover{{border-color:var(--blue)}} .card.active{{border-color:var(--blue);background:var(--surface2)}}
  .card .val{{font-size:1.8rem;font-weight:800;line-height:1;margin-bottom:3px}}
  .card .lbl{{color:var(--muted);font-size:11px;text-transform:uppercase;letter-spacing:.06em}}
  .c-green .val{{color:var(--green)}} .c-red .val{{color:var(--red)}}
  .c-yellow .val{{color:var(--yellow)}} .c-blue .val{{color:var(--blue)}}
  .c-orange .val{{color:var(--orange)}} .c-purple .val{{color:var(--purple)}}
  .c-white .val{{color:var(--text)}}

  .charts{{display:grid;grid-template-columns:300px 1fr;gap:16px;margin-bottom:22px}}
  @media(max-width:860px){{.charts{{grid-template-columns:1fr}}}}
  .chart-box{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:16px}}
  .chart-box h2{{font-size:.8rem;font-weight:600;margin-bottom:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.06em}}
  .chart-box canvas{{max-height:250px}}

  .filters{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);padding:12px 16px;margin-bottom:12px;display:flex;flex-wrap:wrap;gap:10px;align-items:center}}
  .filters label{{color:var(--muted);font-size:12px;margin-right:3px}}
  .filters input,.filters select{{background:var(--surface2);border:1px solid var(--border);color:var(--text);padding:6px 10px;border-radius:6px;font-size:13px;outline:none}}
  .filters input{{width:220px}}
  .filters input:focus,.filters select:focus{{border-color:var(--blue)}}
  .info{{margin-left:auto;color:var(--muted);font-size:12px;white-space:nowrap}}

  .export-row{{display:flex;gap:8px;margin-bottom:10px;flex-wrap:wrap;align-items:center}}
  .export-row span{{color:var(--muted);font-size:12px}}

  .btn{{background:var(--blue);color:#fff;border:none;padding:7px 14px;border-radius:6px;cursor:pointer;font-size:13px;font-weight:600;white-space:nowrap}}
  .btn:hover{{background:#2563eb}}
  .btn-green{{background:#16a34a}} .btn-green:hover{{background:#15803d}}
  .btn-ghost{{background:transparent;border:1px solid var(--border);color:var(--muted)}}
  .btn-ghost:hover{{border-color:var(--blue);color:var(--text)}}

  .table-wrap{{background:var(--surface);border:1px solid var(--border);border-radius:var(--radius);overflow:auto}}
  table{{width:100%;border-collapse:collapse}}
  thead th{{background:var(--surface2);padding:9px 12px;text-align:left;font-size:11px;text-transform:uppercase;letter-spacing:.06em;color:var(--muted);border-bottom:1px solid var(--border);white-space:nowrap;cursor:pointer;user-select:none}}
  thead th:hover{{color:var(--text)}}
  tbody tr{{border-bottom:1px solid var(--border)}}
  tbody tr:hover{{background:var(--surface2)}}
  td{{padding:8px 12px;vertical-align:middle;max-width:280px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}}
  .td-load{{text-align:right;font-variant-numeric:tabular-nums;color:var(--muted);font-size:12px}}
  .load-slow{{color:#f87171}}
  .load-med{{color:#fbbf24}}
  .load-ok{{color:#4ade80}}

  .row-redirect{{background:rgba(245,158,11,.04)}}
  .row-broken{{background:rgba(239,68,68,.07)}}
  .row-warn{{background:rgba(249,115,22,.05)}}
  .row-servererr{{background:rgba(139,92,246,.05)}}
  .row-unkn{{background:rgba(148,163,184,.04)}}

  .pill{{display:inline-block;padding:2px 8px;border-radius:12px;font-size:12px;font-weight:700}}
  .p-ok{{background:rgba(34,197,94,.15);color:#4ade80}}
  .p-redirect{{background:rgba(245,158,11,.15);color:#fbbf24}}
  .p-broken{{background:rgba(239,68,68,.2);color:#f87171}}
  .p-warn{{background:rgba(249,115,22,.15);color:#fb923c}}
  .p-servererr{{background:rgba(139,92,246,.15);color:#a78bfa}}
  .p-unkn{{background:rgba(148,163,184,.15);color:#94a3b8}}

  .tt{{display:inline-block;padding:2px 6px;border-radius:4px;font-size:11px;font-weight:600}}
  .tt-internal{{background:rgba(59,130,246,.15);color:#93c5fd}}
  .tt-external{{background:rgba(139,92,246,.15);color:#c4b5fd}}
  .tt-page{{background:rgba(148,163,184,.15);color:#94a3b8}}

  .bdg{{display:inline-block;padding:2px 6px;border-radius:4px;font-size:11px;font-weight:600}}
  .b-none{{background:rgba(34,197,94,.1);color:#4ade80}}
  .b-low{{background:rgba(245,158,11,.1);color:#fbbf24}}
  .b-lowmed{{background:rgba(249,115,22,.1);color:#fb923c}}
  .b-med{{background:rgba(139,92,246,.1);color:#a78bfa}}
  .b-high{{background:rgba(239,68,68,.15);color:#f87171}}

  .pagination{{display:flex;align-items:center;gap:6px;padding:10px 14px;border-top:1px solid var(--border);flex-wrap:wrap}}
  .pagination button{{background:var(--surface2);border:1px solid var(--border);color:var(--text);padding:4px 10px;border-radius:5px;cursor:pointer;font-size:12px}}
  .pagination button:hover{{border-color:var(--blue)}}
  .pagination button:disabled{{opacity:.35;cursor:not-allowed}}
  .pagination button.active{{background:var(--blue);border-color:var(--blue);color:#fff}}
  .pginfo{{color:var(--muted);font-size:12px;margin:0 4px}}
  .sort-asc::after{{content:" ▲";opacity:.8;font-size:9px}}
  .sort-desc::after{{content:" ▼";opacity:.8;font-size:9px}}
</style>
</head>
<body>

<div class="header">
  <div>
    <h1>Broken Link Report — <span>{base_url}</span>{mode_badge}</h1>
    <div class="meta">
      <span>Run date: <b>{run_date}</b></span>
      <span>Duration: <b>{run_dur}</b></span>
      {"<span>Pages scanned: <b>" + str(target_count) + "</b></span>" if scan_mode == "targeted" else f'<span>Depth: <b>{CONFIG["MAX_DEPTH"]}</b></span>'}
      <span>Workers: <b>{CONFIG["CONCURRENCY"]}</b></span>
      {"<span>Label: <b>" + scan_label + "</b></span>" if scan_label else ""}
    </div>
  </div>
  <div class="hdr-btns">
    <button class="btn btn-green" onclick="exportExcelFull()">⬇ Full Excel</button>
    <button class="btn" onclick="exportCsvFull()">⬇ Full CSV</button>
  </div>
</div>

<div class="container">

{targeted_info_html}
  <div class="cards">
    <div class="card c-white"  onclick="cardFilter('')"          id="card-all">     <div class="val">{total_pages:,}</div><div class="lbl">Parent Pages</div></div>
    <div class="card c-white"  onclick="cardFilter('')"          id="card-links">   <div class="val">{total_links:,}</div><div class="lbl">Total Links</div></div>
    <div class="card c-green"  onclick="cardFilter('ok')"        id="card-ok">      <div class="val">{ok_links:,}</div>  <div class="lbl">200 OK</div></div>
    <div class="card c-yellow" onclick="cardFilter('redirect')"  id="card-redirect"><div class="val">{redirects:,}</div> <div class="lbl">Redirects (3xx)</div></div>
    <div class="card c-red"    onclick="cardFilter('broken')"    id="card-broken">  <div class="val">{broken:,}</div>   <div class="lbl">Broken (404/410)</div></div>
    <div class="card c-purple" onclick="cardFilter('servererr')" id="card-5xx">     <div class="val">{server_errors:,}</div><div class="lbl">Server Errors (5xx)</div></div>
    <div class="card c-orange" onclick="cardFilter('unkn')"      id="card-err">     <div class="val">{errors:,}</div>   <div class="lbl">Timeout / Error</div></div>
    <div class="card c-blue"   onclick="typeFilter2('internal')" id="card-int">     <div class="val">{internal:,}</div> <div class="lbl">Internal Links</div></div>
    <div class="card c-white"  onclick="typeFilter2('external')" id="card-ext">     <div class="val">{external:,}</div> <div class="lbl">External Links</div></div>
  </div>

  <div style="background:#1e2a3f;border:1px solid #3b82f6;border-radius:8px;padding:10px 16px;margin-bottom:16px;font-size:13px;color:#93c5fd;display:flex;align-items:center;gap:12px;flex-wrap:wrap">
    <span>ℹ️  <b>200 OK links are included</b> — useful for tracking external URLs that work today but may break if the destination site changes paths without redirects.</span>
    <label style="display:flex;align-items:center;gap:6px;margin-left:auto;cursor:pointer;color:#e2e8f0;white-space:nowrap">
      <input type="checkbox" id="hide200" onchange="applyFilters()" style="width:14px;height:14px">
      Hide 200 OK rows
    </label>
  </div>

  <div class="charts">
    <div class="chart-box"><h2>Status Distribution</h2><canvas id="donut"></canvas></div>
    <div class="chart-box"><h2>Top Parent Pages by Broken / Error Links</h2><canvas id="bar"></canvas></div>
  </div>

  <div class="filters">
    <div><label>Search</label>
      <input type="text" id="search" placeholder="Filter by URL or anchor text…" oninput="applyFilters()"></div>
    <div><label>Status</label>
      <select id="statusFilter" onchange="applyFilters()">
        <option value="">All Statuses</option>
        <option value="ok">200 OK</option>
        <option value="redirect">Redirect (3xx)</option>
        <option value="broken">Broken (404/410)</option>
        <option value="warn">4xx Other</option>
        <option value="servererr">5xx Server Error</option>
        <option value="unkn">Timeout / Error</option>
      </select></div>
    <div><label>Type</label>
      <select id="typeFilter" onchange="applyFilters()">
        <option value="">All Types</option>
        <option value="Internal">Internal</option>
        <option value="External">External</option>
      </select></div>
    <div><label>Domain</label>
      <input type="text" id="domainFilter" placeholder="e.g. x.com" oninput="applyFilters()" style="width:160px"></div>
    <div><label>Effort</label>
      <select id="effortFilter" onchange="applyFilters()">
        <option value="">All Efforts</option>
        <option value="High">High</option>
        <option value="Medium">Medium</option>
        <option value="Low">Low</option>
        <option value="None">None</option>
      </select></div>
    <button class="btn btn-ghost" onclick="resetFilters()">Reset</button>
    <span class="info" id="rowInfo"></span>
  </div>

  <div class="export-row">
    <span>Export current filter view:</span>
    <button class="btn btn-green" onclick="exportExcelFiltered()">⬇ Filtered Excel</button>
    <button class="btn" onclick="exportCsvFiltered()">⬇ Filtered CSV</button>
    <span class="info" id="filteredCount"></span>
  </div>

  <div class="table-wrap">
    <table>
      <thead id="thead">
        <tr>
          <th onclick="sortBy(0)">Parent URL</th>
          <th onclick="sortBy(1)">Link URL</th>
          <th onclick="sortBy(2)">Anchor Text</th>
          <th onclick="sortBy(3)">Type</th>
          <th onclick="sortBy(4)">Status</th>
          <th onclick="sortBy(5)">Category</th>
          <th onclick="sortBy(6)">Depth</th>
          <th onclick="sortBy(7)">Effort</th>
          <th onclick="sortBy(8)">Load Time</th>
          <th onclick="sortBy(9)" title="IP of the parent page being crawled">Parent IP</th>
          <th onclick="sortBy(10)" title="IP of the link URL">Link IP</th>
          <th onclick="sortBy(11)">Timestamp</th>
        </tr>
      </thead>
      <tbody id="tbody"></tbody>
    </table>
    <div class="pagination" id="pager"></div>
  </div>

</div>

<script>
// Data is base64-encoded JSON — immune to any content in crawled pages.
// No URL, anchor text, or page content can ever break this script block.
let ALL_DATA;
try {{
  ALL_DATA = JSON.parse(atob('{all_data_b64}'));
}} catch(e) {{
  ALL_DATA = [];
  console.error("Failed to decode report data:", e);
  document.addEventListener("DOMContentLoaded", () => {{
    const tb = document.getElementById("tbody");
    if(tb) tb.innerHTML = '<tr><td colspan="12" style="color:#f87171;padding:20px;text-align:center">⚠️ Report data failed to decode. Download the CSV or Excel file for full results.</td></tr>';
  }});
}}
const PAGE_SIZE = {PAGE_SIZE};
let filtered = ALL_DATA.slice();
let sortCol  = -1, sortAsc = true, curPage = 0;

// ── CHARTS ──────────────────────────────────────────────────
// Wrapped in try-catch: CDN failure won't crash renderPage
try {{
  new Chart(document.getElementById('donut').getContext('2d'), {{
    type:'doughnut',
    data:{{labels:{chart_labels},datasets:[{{data:{chart_values},backgroundColor:{chart_colors},borderWidth:2,borderColor:'#1e293b'}}]}},
    options:{{responsive:true,plugins:{{
      legend:{{position:'bottom',labels:{{color:'#94a3b8',padding:10,font:{{size:11}}}}}},
      tooltip:{{callbacks:{{label:ctx=>` ${{ctx.label}}: ${{ctx.parsed.toLocaleString()}}`}}}}
    }}}}
  }});
  new Chart(document.getElementById('bar').getContext('2d'), {{
    type:'bar',
    data:{{labels:{bar_labels},datasets:[{{label:'Issues',data:{bar_values},backgroundColor:'rgba(239,68,68,0.7)',borderRadius:3}}]}},
    options:{{indexAxis:'y',responsive:true,
      plugins:{{legend:{{display:false}},tooltip:{{callbacks:{{label:ctx=>` ${{ctx.parsed.x}} issues`}}}}}},
      scales:{{
        x:{{grid:{{color:'#334155'}},ticks:{{color:'#94a3b8'}}}},
        y:{{grid:{{display:false}},ticks:{{color:'#94a3b8',font:{{size:11}}}}}}
      }}
    }}
  }});
}} catch(e) {{
  console.warn('Chart.js failed to load:', e);
  document.querySelectorAll('.chart-box').forEach(el => {{
    el.innerHTML = '<p style="color:#64748b;padding:20px;text-align:center">Charts unavailable — CDN may be blocked</p>';
  }});
}}

// ── HELPERS ──────────────────────────────────────────────────
function esc(s){{return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}}
function short(s,n){{return s.length>n?s.slice(0,n)+'…':s}}
function loadCell(lm){{
  if(lm===-1||lm===null)return '<td class="td-load">—</td>';
  const cls = lm>3000?'load-slow':lm>1000?'load-med':'load-ok';
  const txt = lm<1000?lm+' ms':(lm/1000).toFixed(1)+'s';
  return `<td class="td-load ${{cls}}">${{txt}}</td>`;
}}
function statusPill(rc,st){{return `<span class="pill p-${{rc}}">${{st}}</span>`}}
function typeBadge(tp){{const c=tp.toLowerCase();return `<span class="tt tt-${{c}}">${{tp}}</span>`}}
function effortBadge(ef){{
  const m={{'None':'b-none','Low':'b-low','Low-Medium':'b-lowmed','Medium':'b-med','High':'b-high'}};
  return `<span class="bdg ${{m[ef]||'b-med'}}">${{ef}}</span>`;
}}

// ── RENDER PAGE ──────────────────────────────────────────────
function renderPage(page){{
  curPage = page;
  const slice = filtered.slice(page*PAGE_SIZE, (page+1)*PAGE_SIZE);
  document.getElementById('tbody').innerHTML = slice.map(r=>`
    <tr class="row-${{r.rc}}">
      <td title="${{esc(r.pu)}}"><a href="${{esc(r.pu)}}" target="_blank">${{esc(short(r.pu,65))}}</a></td>
      <td title="${{esc(r.lu)}}"><a href="${{esc(r.lu)}}" target="_blank">${{esc(short(r.lu,65))}}</a></td>
      <td title="${{esc(r.lt)}}">${{esc(short(r.lt,45))}}</td>
      <td>${{typeBadge(r.tp||'Page')}}</td>
      <td>${{statusPill(r.rc,r.st)}}</td>
      <td>${{esc(r.ca)}}</td>
      <td>${{esc(r.dp)}}</td>
      <td>${{effortBadge(r.ef)}}</td>
      ${{loadCell(r.lm)}}
      <td style="font-family:monospace;font-size:11px;color:#7dd3fc;white-space:nowrap">${{esc(r.pi||'N/A')}}</td>
      <td style="font-family:monospace;font-size:11px;color:#93c5fd;white-space:nowrap">${{esc(r.li||'N/A')}}</td>
      <td>${{esc(r.ts)}}</td>
    </tr>`).join('');
  renderPager();
  const total = filtered.length;
  document.getElementById('rowInfo').textContent =
    `${{total.toLocaleString()}} rows · page ${{page+1}} of ${{Math.max(1,Math.ceil(total/PAGE_SIZE))}}`;
  document.getElementById('filteredCount').textContent =
    `${{total.toLocaleString()}} rows in current filter`;
}}

// ── PAGINATOR ────────────────────────────────────────────────
function renderPager(){{
  const total=Math.ceil(filtered.length/PAGE_SIZE);
  const pager=document.getElementById('pager');
  if(total<=1){{pager.innerHTML='';return;}}
  let pages=[];
  if(total<=9){{pages=Array.from({{length:total}},(_,i)=>i);}}
  else{{
    pages=[0];
    let lo=Math.max(1,curPage-3),hi=Math.min(total-2,curPage+3);
    if(lo>1)pages.push('…');
    for(let i=lo;i<=hi;i++)pages.push(i);
    if(hi<total-2)pages.push('…');
    pages.push(total-1);
  }}
  pager.innerHTML=
    `<button onclick="renderPage(${{curPage-1}})" ${{curPage===0?'disabled':''}}>‹ Prev</button>`+
    pages.map(p=>p==='…'?`<span class="pginfo">…</span>`:
      `<button class="${{p===curPage?'active':''}}" onclick="renderPage(${{p}})">${{p+1}}</button>`
    ).join('')+
    `<button onclick="renderPage(${{curPage+1}})" ${{curPage>=total-1?'disabled':''}}>Next ›</button>`;
}}

// ── FILTER ───────────────────────────────────────────────────
let _searchTimer=null;
function applyFilters(){{clearTimeout(_searchTimer);_searchTimer=setTimeout(_applyFilters,120);}}
function _applyFilters(){{
  const q=document.getElementById('search').value.toLowerCase().trim();
  const st=document.getElementById('statusFilter').value;
  const tp=document.getElementById('typeFilter').value;
  const ef=document.getElementById('effortFilter').value;
  const hide200=document.getElementById('hide200')?.checked;
  const dom=(document.getElementById('domainFilter')?.value||'').toLowerCase().trim();
  filtered=ALL_DATA.filter(r=>{{
    if(hide200&&r.rc==='ok')return false;
    if(st&&r.rc!==st)return false;
    if(tp&&r.tp!==tp)return false;
    if(ef&&r.ef!==ef)return false;
    if(dom&&!r.lu.toLowerCase().includes(dom))return false;
    if(q&&!(r.pu.toLowerCase().includes(q)||r.lu.toLowerCase().includes(q)||r.lt.toLowerCase().includes(q)||(r.pi||'').includes(q)||(r.li||'').includes(q)))return false;
    return true;
  }});
  if(sortCol>=0)sortFiltered();
  renderPage(0);
  updateCardHighlight();
}}
function resetFilters(){{
  document.getElementById('search').value='';
  document.getElementById('statusFilter').value='';
  document.getElementById('typeFilter').value='';
  document.getElementById('effortFilter').value='';
  const h200=document.getElementById('hide200');if(h200)h200.checked=false;
  const hdom=document.getElementById('domainFilter');if(hdom)hdom.value='';
  filtered=ALL_DATA.slice();sortCol=-1;sortAsc=true;
  document.querySelectorAll('#thead th').forEach(th=>th.classList.remove('sort-asc','sort-desc'));
  document.querySelectorAll('.card').forEach(c=>c.classList.remove('active'));
  renderPage(0);
}}
function cardFilter(rc){{
  document.getElementById('statusFilter').value=rc;
  document.getElementById('typeFilter').value='';
  document.getElementById('effortFilter').value='';
  document.getElementById('search').value='';
  _applyFilters();
}}
function typeFilter2(tp){{
  document.getElementById('typeFilter').value=tp.charAt(0).toUpperCase()+tp.slice(1);
  document.getElementById('statusFilter').value='';
  document.getElementById('effortFilter').value='';
  document.getElementById('search').value='';
  _applyFilters();
}}
function updateCardHighlight(){{
  const st=document.getElementById('statusFilter').value;
  const tp=document.getElementById('typeFilter').value.toLowerCase();
  document.querySelectorAll('.card').forEach(c=>c.classList.remove('active'));
  const m={{'ok':'card-ok','redirect':'card-redirect','broken':'card-broken','servererr':'card-5xx','unkn':'card-err'}};
  if(m[st])document.getElementById(m[st])?.classList.add('active');
  if(tp==='internal')document.getElementById('card-int')?.classList.add('active');
  if(tp==='external')document.getElementById('card-ext')?.classList.add('active');
}}

// ── SORT ─────────────────────────────────────────────────────
const SORT_KEYS=['pu','lu','lt','tp','st','ca','dp','ef','lm','pi','li','ts'];
function sortBy(col){{
  if(sortCol===col){{sortAsc=!sortAsc;}}else{{sortCol=col;sortAsc=true;}}
  document.querySelectorAll('#thead th').forEach((th,i)=>{{
    th.classList.remove('sort-asc','sort-desc');
    if(i===col)th.classList.add(sortAsc?'sort-asc':'sort-desc');
  }});
  sortFiltered();renderPage(0);
}}
function sortFiltered(){{
  const key=SORT_KEYS[sortCol];
  filtered.sort((a,b)=>{{
    const va=String(a[key]??''),vb=String(b[key]??'');
    const n=Number(va)-Number(vb);
    const cmp=isNaN(n)?va.localeCompare(vb):n;
    return sortAsc?cmp:-cmp;
  }});
}}

// ── EXCEL EXPORT (SheetJS) ───────────────────────────────────
const XL_HEADERS = ['Parent URL','Link URL','Anchor Text','Type','Status',
                    'Category','Depth','Effort','Load Time (ms)','Timestamp'];
const XL_KEYS    = ['pu','lu','lt','tp','st','ca','dp','ef','lm','ts'];

function rowsToSheet(data){{
  const ws_data = [XL_HEADERS, ...data.map(r=>XL_KEYS.map(k=>{{
    const v=r[k]??'';
    // Load time: store as number, -1 becomes blank
    if(k==='lm') return v===-1?'':Number(v);
    // Depth: numeric
    if(k==='dp') return v===''?'':Number(v)||v;
    return String(v);
  }}))];
  const ws = XLSX.utils.aoa_to_sheet(ws_data);
  // Column widths
  ws['!cols']=[{{wch:60}},{{wch:60}},{{wch:35}},{{wch:10}},{{wch:12}},
               {{wch:18}},{{wch:7}},{{wch:12}},{{wch:14}},{{wch:24}}];
  return ws;
}}

function buildWorkbook(data, sheetName){{
  const wb = XLSX.utils.book_new();
  const ws = rowsToSheet(data);
  XLSX.utils.book_append_sheet(wb, ws, sheetName);

  // Summary sheet
  const now_str = new Date().toLocaleString('en-US',{{timeZone:'America/New_York'}});
  const summary = [
    ['Broken Link Report'],
    ['Site', '{base_url}'],
    ['Run Date (ET)',  '{run_date}'],
    ['Duration', '{run_dur}'],
    [''],
    ['Metric','Count'],
    ['Parent Pages', {total_pages}],
    ['Total Links', {total_links}],
    ['200 OK', {ok_links}],
    ['Redirects (3xx)', {redirects}],
    ['Broken (404/410)', {broken}],
    ['Server Errors (5xx)', {server_errors}],
    ['Timeout / Error', {errors}],
    ['Internal Links', {internal}],
    ['External Links', {external}],
  ];
  const ws2 = XLSX.utils.aoa_to_sheet(summary);
  ws2['!cols']=[{{wch:24}},{{wch:40}}];
  XLSX.utils.book_append_sheet(wb, ws2, 'Summary');
  return wb;
}}

function exportExcelFull(){{
  const wb = buildWorkbook(ALL_DATA, 'All Links');
  XLSX.writeFile(wb, 'broken_links_full_report.xlsx');
}}

function exportExcelFiltered(){{
  const wb = buildWorkbook(filtered, 'Filtered Results');
  XLSX.writeFile(wb, 'broken_links_filtered.xlsx');
}}

// ── CSV EXPORT ───────────────────────────────────────────────
function rowsToCsv(data){{
  const headers = ['Parent URL','Link URL','Anchor Text','Type','Status',
                   'Category','Depth','Effort','Load Time (ms)','Timestamp'];
  const keys    = ['pu','lu','lt','tp','st','ca','dp','ef','lm','ts'];
  const escape  = v => {{
    const s = v===null||v===undefined?'':String(v);
    return s.includes(',') || s.includes('"') || s.includes('\\n')
      ? '"' + s.replace(/"/g,'""') + '"' : s;
  }};
  const rows = [headers.join(',')];
  data.forEach(r => {{
    rows.push(keys.map(k => {{
      const v = r[k]??'';
      return escape(k==='lm'&&v===-1?'':v);
    }}).join(','));
  }});
  return rows.join('\\n');
}}

function downloadCsv(content, filename){{
  const blob = new Blob(['\uFEFF' + content], {{type:'text/csv;charset=utf-8;'}});
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url; a.download = filename; a.click();
  URL.revokeObjectURL(url);
}}

function exportCsvFull(){{
  downloadCsv(rowsToCsv(ALL_DATA), 'broken_links_full_report.csv');
}}

function exportCsvFiltered(){{
  downloadCsv(rowsToCsv(filtered), 'broken_links_filtered.csv');
}}

// ── INIT ─────────────────────────────────────────────────────
renderPage(0);
</script>
</body>
</html>"""
    return html

# ──────────────────────────────────────────────
#  EXCEL OUTPUT  (openpyxl)
# ──────────────────────────────────────────────

def write_excel(results: list[dict], path: str):
    """Write a formatted .xlsx file with two sheets: All Links + Summary."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning("openpyxl not installed — skipping Excel output")
        return

    STATUS_FILLS = {
        "ok":        PatternFill("solid", fgColor="D4EDDA"),
        "redirect":  PatternFill("solid", fgColor="FFF3CD"),
        "broken":    PatternFill("solid", fgColor="F8D7DA"),
        "warn":      PatternFill("solid", fgColor="FFE5CC"),
        "servererr": PatternFill("solid", fgColor="E2D9F3"),
        "unkn":      PatternFill("solid", fgColor="E9ECEF"),
    }
    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
    BOLD      = Font(bold=True)
    thin      = Side(border_style="thin", color="CCCCCC")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    def rc(status):
        s = str(status)
        if s == "200":         return "ok"
        if s.startswith("3"):  return "redirect"
        if s in ("404","410"): return "broken"
        if s.startswith("4"):  return "warn"
        if s.startswith("5"):  return "servererr"
        return "unkn"

    wb = openpyxl.Workbook()

    EXCEL_MAX_ROWS = 1_048_576  # Excel hard limit including header row
    headers    = ["Parent URL","Link URL","Anchor Text","Type","Status",
                  "Category","Depth","Effort","Load Time (ms)","Timestamp",
                  "Parent IP","Link IP"]
    col_widths = [70, 70, 35, 12, 12, 18, 8, 14, 16, 24, 16, 16]

    def make_data_sheet(wb, title, rows, is_first=False):
        """Write up to EXCEL_MAX_ROWS-1 data rows to a sheet with headers."""
        ws = wb.active if is_first else wb.create_sheet(title)
        ws.title = title
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = BORDER
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 18
        ws.freeze_panes = "A2"
        for ri, r in enumerate(rows, 2):
            if ri > EXCEL_MAX_ROWS:  # safety — never exceed Excel limit
                log.warning("Excel sheet '%s' capped at %d rows", title, EXCEL_MAX_ROWS - 1)
                break
            lm = r.get("load_ms", -1)
            load_val = lm if lm not in (None, -1) else None
            row_vals = [
                r.get("page_url",""), r.get("link_url",""),
                r.get("link_text",""), r.get("link_type",""),
                str(r.get("status","")), r.get("category",""),
                r.get("depth",""), r.get("effort",""),
                load_val, str(r.get("timestamp","")),
                r.get("page_ip","N/A"), r.get("link_ip","N/A"),
            ]
            fill = STATUS_FILLS.get(rc(r.get("status","")))
            for ci, val in enumerate(row_vals, 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border    = BORDER
                cell.alignment = Alignment(vertical="center",
                                           wrap_text=(ci in (1,2)))
                if fill:
                    cell.fill = fill
        ws.auto_filter.ref = ws.dimensions
        return ws

    # ── Sheet 1+: All Links (split into chunks if > Excel row limit) ─────────
    rows_per_sheet = EXCEL_MAX_ROWS - 1  # subtract header row
    total_rows     = len(results)
    num_sheets     = max(1, -(-total_rows // rows_per_sheet))  # ceiling division

    if num_sheets == 1:
        make_data_sheet(wb, "All Links", results, is_first=True)
    else:
        log.warning(
            "Data has %d rows — exceeds Excel limit (%d). "
            "Splitting into %d sheets. Full data is in the CSV.",
            total_rows, EXCEL_MAX_ROWS, num_sheets)
        for i in range(num_sheets):
            chunk = results[i * rows_per_sheet : (i + 1) * rows_per_sheet]
            title = f"All Links {i+1}" if i > 0 else "All Links"
            make_data_sheet(wb, title, chunk, is_first=(i == 0))

    # ── Sheet 2: Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    broken = sum(1 for r in results if str(r.get("status","")) in ("404","410","451"))
    ok     = sum(1 for r in results if str(r.get("status","")) == "200")
    redirects = sum(1 for r in results if str(r.get("status","")).startswith("3"))
    server_errors = sum(1 for r in results if str(r.get("status","")).startswith("5"))
    errors = sum(1 for r in results if "Error" in str(r.get("status","")) or "Timeout" in str(r.get("status","")))
    internal = sum(1 for r in results if r.get("link_type") == "Internal")
    external = sum(1 for r in results if r.get("link_type") == "External")
    pages    = len({r["page_url"] for r in results})

    summary_rows = [
        ("Broken Link Report", None),
        ("Site", CONFIG["BASE_URL"]),
        ("Run Date (ET)",  now_est().strftime("%Y-%m-%d %H:%M:%S %Z")),
        (None, None),
        ("Metric", "Count"),
        ("Parent Pages Crawled", pages),
        ("Total Links Found", len(results)),
        ("200 OK", ok),
        ("Redirects (3xx)", redirects),
        ("Broken (404/410)", broken),
        ("Server Errors (5xx)", server_errors),
        ("Timeout / Errors", errors),
        ("Internal Links", internal),
        ("External Links", external),
    ]
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 45
    for ri2, (label, value) in enumerate(summary_rows, 1):
        if label:
            c = ws2.cell(row=ri2, column=1, value=label)
            if ri2 == 1:
                c.font = Font(bold=True, size=14, color="1F4E79")
            elif label == "Metric":
                c.font = HDR_FONT
                c.fill = HDR_FILL
                vc = ws2.cell(row=ri2, column=2, value=value)
                vc.font = HDR_FONT
                vc.fill = HDR_FILL
                continue
            else:
                c.font = BOLD
        if value is not None:
            ws2.cell(row=ri2, column=2, value=value)

    wb.save(path)
    log.info("Excel written → %s  (%d rows)", path, len(results))


async def main():
    start_time = time.time()
    log.info("Starting — BASE_URL=%s", CONFIG["BASE_URL"])
    pages_display = "unlimited" if CONFIG["MAX_PAGES"] == 0 else str(CONFIG["MAX_PAGES"])
    log.info("Config: MAX_DEPTH=%d  MAX_PAGES=%s  MAX_QUEUE=%d  CONCURRENCY=%d  TIMEOUT=%ds",
             CONFIG["MAX_DEPTH"], pages_display, CONFIG["MAX_QUEUE"],
             CONFIG["CONCURRENCY"], CONFIG["TIMEOUT"])

    # ── Decide crawl mode ────────────────────────────────────────
    target_pages = CONFIG.get("TARGET_PAGES", [])
    scan_mode    = "targeted" if target_pages else "full"

    if scan_mode == "targeted":
        log.info("Mode: TARGETED — %d page(s) specified", len(target_pages))
        results = await targeted_crawl(target_pages)
        checkpoint_saved = False
    else:
        log.info("Mode: FULL CRAWL — starting from %s", CONFIG["BASE_URL"])
        results, checkpoint_saved = await crawl()

    output_dir = Path(CONFIG["OUTPUT_DIR"])
    output_dir.mkdir(parents=True, exist_ok=True)

    if not results:
        log.warning("No results collected. Check URLs and network access.")
        # Write an empty index.html so GitHub Pages + Verify step don't fail
        empty_html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<title>Crawl Report — No Results</title>
<style>body{{background:#0f172a;color:#e2e8f0;font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;margin:0}}.box{{text-align:center;padding:40px;background:#1e293b;border-radius:12px;border:1px solid #334155}}</style>
</head><body><div class="box">
<h1 style="color:#f59e0b">⚠️ No Results</h1>
<p>The crawler returned no data.</p>
<p style="color:#94a3b8">Check that <b>{CONFIG["BASE_URL"]}</b> is reachable and the secret is set correctly.</p>
</div></body></html>"""
        (output_dir / "index.html").write_text(empty_html, encoding="utf-8")
        import json as _json
        _json.dump({"url": CONFIG["BASE_URL"], "mode": "failed", "label": "",
                     "run_date": now_est().strftime("%Y-%m-%d %H:%M:%S %Z"),
                     "duration": "0s", "pages": 0, "total_links": 0, "ok": 0,
                     "broken": 0, "redirects": 0, "server_errors": 0,
                     "timeouts": 0, "internal": 0, "external": 0},
                    open(str(output_dir / "summary.json"), "w"))
        return

    elapsed  = time.time() - start_time
    dur_str  = fmt_duration(elapsed)
    run_date = now_est().strftime("%Y-%m-%d %H:%M:%S %Z")
    ts_tag   = now_est().strftime("%Y%m%d_%H%M%S")
    label    = CONFIG.get("SCAN_LABEL", "")
    safe_label = label.replace(" ", "_").replace("/", "_")[:30] if label else ""
    file_tag   = f"{ts_tag}_{safe_label}" if safe_label else ts_tag

    csv_path   = str(output_dir / f"broken_links_{file_tag}.csv")
    html_path  = str(output_dir / f"report_{file_tag}.html")
    xlsx_path  = str(output_dir / f"broken_links_{file_tag}.xlsx")

    # Resolve IPs for parent pages and link URLs (one DNS call per unique hostname)
    await enrich_with_ips(results)

    write_csv(results, csv_path)
    # Excel: write non-200 rows only (broken/redirect/error/timeout).
    # The HTML report now includes 200 OK rows with a hide/show toggle.
    # Excel is kept as the actionable broken-links list to stay under size limits.
    # All rows including 200 OK are available in the CSV for full audits.
    excel_rows = [r for r in results if str(r.get("status","")) != "200"]
    log.info("Excel: writing %d non-200 rows (%d OK rows in CSV only)",
             len(excel_rows), len(results) - len(excel_rows))
    write_excel(excel_rows, xlsx_path)

    html = build_html_report(results, csv_path, elapsed,
                             scan_mode=scan_mode,
                             target_pages=target_pages)
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info("HTML report written → %s", html_path)

    # index.html — GitHub Pages always serves the most recent run
    index_path = str(output_dir / "index.html")
    with open(index_path, "w", encoding="utf-8") as f:
        f.write(html)
    log.info("index.html written → %s (used by GitHub Pages)", index_path)

    broken = sum(1 for r in results if str(r["status"]) in ("404","410","451"))
    ok     = sum(1 for r in results if str(r["status"]) == "200")
    mode_str = f"TARGETED ({len(target_pages)} pages)" if scan_mode == "targeted" else "FULL CRAWL"
    print(f"\n{'─'*60}")
    print(f"  SCAN SUMMARY  [{mode_str}]")
    if label: print(f"  Label      : {label}")
    print(f"  Base URL   : {CONFIG['BASE_URL']}")
    print(f"  Run Date   : {run_date}")
    pages_crawled = len({r['page_url'] for r in results})
    print(f"  Pages      : {pages_crawled:,}")
    print(f"  Total links: {len(results):,}")
    print(f"  200 OK     : {ok:,}")
    print(f"  Broken     : {broken:,}")
    print(f"  Duration   : {dur_str}")
    print(f"  Reports    : {csv_path}")
    print(f"              {html_path}")
    print(f"              {xlsx_path}")
    print(f"{'─'*60}\n")

    # Write summary.json so the workflow can read stats for the email body
    redirects  = sum(1 for r in results if str(r["status"]).startswith("3"))
    srv_errors = sum(1 for r in results if str(r["status"]).startswith("5"))
    timeouts   = sum(1 for r in results if "Error" in str(r["status"]) or "Timeout" in str(r["status"]))
    internal   = sum(1 for r in results if r.get("link_type") == "Internal")
    external   = sum(1 for r in results if r.get("link_type") == "External")
    summary_data = {
        "url":           CONFIG["BASE_URL"],
        "mode":          mode_str,
        "label":         label,
        "run_date":      run_date,
        "duration":      dur_str,
        "pages":         pages_crawled,
        "total_links":   len(results),
        "ok":            ok,
        "broken":        broken,
        "redirects":     redirects,
        "server_errors": srv_errors,
        "timeouts":      timeouts,
        "internal":      internal,
        "external":      external,
    }
    summary_data["checkpoint_saved"] = checkpoint_saved
    summary_data["run_count"] = int(os.getenv("RESUME_RUN_COUNT", "1"))
    summary_path = str(output_dir / "summary.json")
    with open(summary_path, "w", encoding="utf-8") as sf:
        json.dump(summary_data, sf, indent=2)
    log.info("summary.json written → %s", summary_path)

    # Write flag file so the workflow knows whether to chain another run
    flag_path = str(output_dir / "checkpoint_saved.txt")
    with open(flag_path, "w") as ff:
        ff.write("true" if checkpoint_saved else "false")
    if checkpoint_saved:
        log.info("🔁 checkpoint_saved.txt = true — workflow will auto-dispatch next run")
    else:
        log.info("✅ checkpoint_saved.txt = false — this was the final run")

    if os.getenv("GITHUB_STEP_SUMMARY"):
        with open(os.environ["GITHUB_STEP_SUMMARY"], "a") as gf:
            gf.write(f"## {'🎯 Targeted Scan' if scan_mode=='targeted' else '🌐 Full Crawl'} — `{CONFIG['BASE_URL']}`\n\n")
            if label:
                gf.write(f"**Label:** {label}\n\n")
            if scan_mode == "targeted":
                gf.write(f"**Pages scanned:**\n")
                for u in target_pages:
                    gf.write(f"- {u}\n")
                gf.write("\n")
            gf.write(f"| Metric | Value |\n|---|---|\n")
            gf.write(f"| Mode | {mode_str} |\n")
            gf.write(f"| Run Date (ET)  | {run_date} |\n")
            gf.write(f"| Pages Scanned | {pages_crawled:,} |\n")
            gf.write(f"| Total Links | {len(results):,} |\n")
            gf.write(f"| 200 OK | {ok:,} |\n")
            gf.write(f"| Broken (404/410) | {broken:,} |\n")
            gf.write(f"| Duration | {dur_str} |\n")


if __name__ == "__main__":
    asyncio.run(main())
